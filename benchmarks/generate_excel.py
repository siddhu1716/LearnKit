import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    excel_path = "/home/mcw/Desktop/personal/benchmarks_comparison.xlsx"
    print(f"Creating Excel file at {excel_path}...")
    
    # ── TAB 1: MASTER COMPARISON ──
    master_data = [
        # Domain | Arm | Tasks Run | Success Metric | Metric Value | Avg Latency (s) | Avg Tokens | Token Savings vs. Raw Logs
        {
            "Benchmark Domain": "PBEBench-Lite (String Trans.)",
            "Arm": "Control",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "95.0%",
            "Avg Latency (s)": 0.44,
            "Avg Tokens": 840.6,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "PBEBench-Lite (String Trans.)",
            "Arm": "Cold Start (LearnKit)",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "100.0%",
            "Avg Latency (s)": 0.31,
            "Avg Tokens": 841.8,
            "Context Token Savings": "85.1%"
        },
        {
            "Benchmark Domain": "PBEBench-Lite (String Trans.)",
            "Arm": "Warmed Start (LearnKit)",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "100.0%",
            "Avg Latency (s)": 0.30,
            "Avg Tokens": 843.5,
            "Context Token Savings": "85.1%"
        },
        {
            "Benchmark Domain": "SLR-Bench (Logic Reasoning)",
            "Arm": "Control",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "100.0%",
            "Avg Latency (s)": 0.38,
            "Avg Tokens": 581.0,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "SLR-Bench (Logic Reasoning)",
            "Arm": "Cold Start (LearnKit)",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "100.0%",
            "Avg Latency (s)": 0.37,
            "Avg Tokens": 797.1,
            "Context Token Savings": "80.5%"
        },
        {
            "Benchmark Domain": "SLR-Bench (Logic Reasoning)",
            "Arm": "Warmed Start (LearnKit)",
            "Tasks Run": 20,
            "Success Metric": "Pass Rate",
            "Metric Value": "100.0%",
            "Avg Latency (s)": 0.36,
            "Avg Tokens": 798.5,
            "Context Token Savings": "80.5%"
        },
        {
            "Benchmark Domain": "Python Debugging",
            "Arm": "Control",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.10 / 5.0",
            "Avg Latency (s)": 0.85,
            "Avg Tokens": 220.0,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "Python Debugging",
            "Arm": "Treatment (LearnKit)",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.40 / 5.0",
            "Avg Latency (s)": 0.92,
            "Avg Tokens": 351.0,
            "Context Token Savings": "78.4%"
        },
        {
            "Benchmark Domain": "SQL Authoring",
            "Arm": "Control",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.50 / 5.0",
            "Avg Latency (s)": 0.78,
            "Avg Tokens": 231.0,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "SQL Authoring",
            "Arm": "Treatment (LearnKit)",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.50 / 5.0",
            "Avg Latency (s)": 0.81,
            "Avg Tokens": 358.0,
            "Context Token Savings": "79.2%"
        },
        {
            "Benchmark Domain": "Contract Summarization",
            "Arm": "Control",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.70 / 5.0",
            "Avg Latency (s)": 1.25,
            "Avg Tokens": 302.0,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "Contract Summarization",
            "Arm": "Treatment (LearnKit)",
            "Tasks Run": 10,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "4.90 / 5.0",
            "Avg Latency (s)": 1.34,
            "Avg Tokens": 605.0,
            "Context Token Savings": "81.0%"
        },
        {
            "Benchmark Domain": "SWE-bench Lite (Pytest)",
            "Arm": "Control (Zero-Shot)",
            "Tasks Run": 3,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "2.00 / 5.0",
            "Avg Latency (s)": 12.4,
            "Avg Tokens": 4000.0,
            "Context Token Savings": "N/A"
        },
        {
            "Benchmark Domain": "SWE-bench Lite (Pytest)",
            "Arm": "Warmed Start (Seeded Skill)",
            "Tasks Run": 1,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "3.00 / 5.0",
            "Avg Latency (s)": 14.1,
            "Avg Tokens": 4300.0,
            "Context Token Savings": "89.8%"
        },
        {
            "Benchmark Domain": "SWE-bench Lite (Pytest)",
            "Arm": "Continuous Evolution (30 runs)",
            "Tasks Run": 30,
            "Success Metric": "Mean Judge Score",
            "Metric Value": "2.33 / 5.0",
            "Avg Latency (s)": 10.8,
            "Avg Tokens": 4250.0,
            "Context Token Savings": "89.8%"
        }
    ]
    df_master = pd.DataFrame(master_data)

    # ── TAB 2: CONTINUOUS EVOLUTION (SWE) ──
    swe_evolution = [
        {"Iteration": 1, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 1},
        {"Iteration": 2, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.0, "Average Score": 2.333, "Fuzzy-Match Errors": 0},
        {"Iteration": 3, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 4, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 5, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 6, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 7, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.0, "Average Score": 2.333, "Fuzzy-Match Errors": 0},
        {"Iteration": 8, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.5, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 9, "pytest-7432": 2.5, "pytest-7490": 2.0, "pytest-8906": 3.0, "Average Score": 2.500, "Fuzzy-Match Errors": 0},
        {"Iteration": 10, "pytest-7432": 2.0, "pytest-7490": 2.0, "pytest-8906": 3.0, "Average Score": 2.333, "Fuzzy-Match Errors": 0},
    ]
    df_swe = pd.DataFrame(swe_evolution)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master Summary", index=False)
        df_swe.to_excel(writer, sheet_name="SWE Evolution", index=False)

    # ── STYLING THE EXCEL WORKBOOK ──
    wb = openpyxl.load_workbook(excel_path)
    
    # Fonts & Colors
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    accent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Style Tab 1
    ws1 = wb["Master Summary"]
    ws1.insert_rows(1, 2)
    ws1["A1"] = "LearnKit Master Benchmark Performance Summary"
    ws1["A1"].font = title_font
    
    # Headers row is now 3
    for col_idx in range(1, 9):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(4, 4 + len(master_data)):
        # Alternate zebra coloring by domain
        domain = ws1.cell(row=row_idx, column=1).value
        # Use soft color grouping for clarity
        bg_fill = zebra_fill if row_idx % 2 == 0 else openpyxl.styles.fills.PatternFill(fill_type=None)
        
        for col_idx in range(1, 9):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if cell.value in ["100.0%", "4.90 / 5.0", "3.00 / 5.0"] or (col_idx == 8 and cell.value != "N/A"):
                cell.fill = accent_fill
            elif bg_fill.fill_type:
                cell.fill = bg_fill
            
            # Alignments
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style Tab 2
    ws2 = wb["SWE Evolution"]
    ws2.insert_rows(1, 2)
    ws2["A1"] = "Pytest Skipping Tasks — 30-Run Continuous Evolution Analysis"
    ws2["A1"].font = title_font
    
    for col_idx in range(1, 7):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(4, 4 + len(swe_evolution)):
        for col_idx in range(1, 7):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Color highlight for evolution points
            if col_idx == 5:
                cell.fill = accent_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row < 3:  # Skip titles
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(excel_path)
    print("Successfully saved and formatted Excel sheet.")

if __name__ == "__main__":
    main()
